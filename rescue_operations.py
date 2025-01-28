import datetime
import openpyxl
from openpyxl import Workbook

class RescueAction:
    def __init__(self, incident_type, vehicles, location, date, weather_conditions, temperature, season):
        self.incident_type = incident_type
        self.vehicles = vehicles
        self.location = location
        self.date = date
        self.weather_conditions = weather_conditions
        self.temperature = temperature
        self.season = season

    def display_info(self):
        print(f"Incident Type: {self.incident_type}")
        print(f"Vehicles: {', '.join(self.vehicles)}")
        print(f"Location: {self.location}")
        print(f"Date: {self.date}")
        print(f"Weather Conditions: {', '.join(self.weather_conditions)}")
        print(f"Temperature: {self.temperature}")
        print(f"Season: {self.season}")


def enter_date():
    while True:
        date_input = input("Enter date (YYYY-MM-DD): ")
        try:
            date = datetime.datetime.strptime(date_input, "%Y-%m-%d").date()
            return date
        except ValueError:
            print("Invalid date format. Please try again.")


def save_to_excel(actions, filename="rescue_actions.xlsx"):
    try:
        # Try to open an existing Excel file
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        # Create a new file if it doesn't exist
        workbook = Workbook()
        sheet = workbook.active
        # Column headers (only if the file is newly created)
        sheet.append(["Incident Type", "Vehicles", "Location", "Date", "Weather Conditions", "Temperature", "Season"])

    for action in actions:
        vehicles = ', '.join(action.vehicles)
        weather_conditions = ', '.join(action.weather_conditions)
        sheet.append([action.incident_type, vehicles, action.location, action.date, weather_conditions, action.temperature, action.season])

    workbook.save(filename)


def load_from_excel(filename="rescue_actions.xlsx"):
    actions = []
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            actions.append(RescueAction(*row))

    except FileNotFoundError:
        print("File does not exist. Creating a new file.")
    return actions


if __name__ == "__main__":
    rescue_actions = load_from_excel()

    try:
        while True:
            print("\n1. Add a new rescue action")
            print("2. Display all rescue actions")
            print("3. Save to Excel file")
            print("4. Exit")

            choice = input("Choose an option: ")

            if choice == "1":
                print("Choose incident type:")
                print("1 - Accident")
                print("2 - Indoor Fire")
                print("3 - Outdoor Fire")
                print("4 - Missing Person Search")
                print("5 - Surveillance Check")
                print("6 - PSP Operational Zone Security")
                print("7 - Training")
                print("8 - Other")

                incident_choice = input("Select incident type number: ")

                if incident_choice == "8":
                    incident_type = input("Enter a different incident type: ")
                    comment = input("Add a comment: ")
                    incident_type += f" ({comment})"
                elif incident_choice in ["1", "2", "3", "4", "5", "6", "7"]:
                    incident_type = {
                        "1": "Accident",
                        "2": "Indoor Fire",
                        "3": "Outdoor Fire",
                        "4": "Missing Person Search",
                        "5": "Surveillance Check",
                        "6": "PSP Operational Zone Security",
                        "7": "Training"
                    }[incident_choice]
                else:
                    print("Invalid choice. Please try again.")
                    continue

                location = input("Enter location: ")
                date = enter_date()

                print("Choose vehicles:")
                print("1 - GLBA 0.4/2")
                print("2 - GBA 2.5/16")
                print("3 - GCBA 5/32")
                print("4 - Ladder (D)")
                print("5 - Hydraulic Lift (H)")

                selected_vehicles = []
                while True:
                    vehicle_choice = input("Select vehicle number (type 'end' when done): ")
                    if vehicle_choice.lower() == "end":
                        break
                    elif vehicle_choice in ["1", "2", "3", "4", "5"]:
                        vehicle_dict = {
                            "1": "GLBA 0.4/2",
                            "2": "GBA 2.5/16",
                            "3": "GCBA 5/32",
                            "4": "Ladder (D)",
                            "5": "Hydraulic Lift (H)"
                        }
                        selected_vehicles.append(vehicle_dict[vehicle_choice])
                    else:
                        print("Invalid choice. Please try again.")

                print("Choose weather conditions:")
                print("1 - Wind")
                print("2 - Rain")
                print("3 - Snow")
                print("4 - Frost")
                print("5 - Normal Conditions")

                selected_conditions = []
                while True:
                    weather_choice = input("Select weather condition number (type 'end' when done): ")
                    if weather_choice.lower() == "end":
                        break
                    elif weather_choice == "5":
                        selected_conditions = ["Normal Conditions"]
                        break
                    elif weather_choice in ["1", "2", "3", "4"]:
                        condition_dict = {
                            "1": "Wind",
                            "2": "Rain",
                            "3": "Snow",
                            "4": "Frost"
                        }
                        selected_conditions.append(condition_dict[weather_choice])
                    else:
                        print("Invalid choice. Please try again.")

                temperature = input("Enter temperature: ")

                print("Summary of entered data:")
                print(f"Incident Type: {incident_type}")
                print(f"Location: {location}")
                print(f"Date: {date}")
                print(f"Vehicles: {', '.join(selected_vehicles)}")
                print(f"Weather Conditions: {', '.join(selected_conditions)}")

                new_action = RescueAction(incident_type, selected_vehicles, location, date, selected_conditions, temperature, "")
                rescue_actions.append(new_action)
                print("New rescue action added.")

            elif choice == "2":
                for action in rescue_actions:
                    action.display_info()

            elif choice == "3":
                save_to_excel(rescue_actions)
                print("Data saved to Excel file.")

            elif choice == "4":
                break

            else:
                print("Invalid choice. Please try again.")

    except KeyboardInterrupt:
        print("\nInterrupted by user. Closing program.")
